"""报表导出服务：生成 Excel 文件"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from sqlalchemy.orm import Session, joinedload

from app.models.product import Product
from app.models.profit_record import ProfitRecord

# 与前端 REGION_MAP 一致，便于阅读
REGION_LABEL_ZH = {
    "PH": "菲律宾",
    "MY": "马来西亚",
    "TH": "泰国",
    "SG": "新加坡",
    "ID": "印尼",
    "VN": "越南",
}

STATUS_LABEL_ZH = {
    "pending": "待定",
    "selected": "已选",
    "abandoned": "放弃",
    "erp_synced": "已同步ERP",
}

STOCK_LABEL_ZH = {
    0: "下架",
    1: "在售",
}


def _fmt_dt(v: Optional[datetime]) -> Optional[str]:
    if v is None:
        return None
    return v.strftime("%Y-%m-%d %H:%M:%S")


def _num(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def _profit_rate_pct(v: Any) -> Optional[str]:
    if v is None:
        return None
    try:
        return f"{float(v) * 100:.2f}%"
    except (TypeError, ValueError):
        return str(v)


def _get_primary_match(product: Product):
    matches = product.pdd_matches or []
    primary = next((m for m in matches if m.is_primary), None)
    return primary or (matches[0] if matches else None)


def ensure_tiktok_url_locale_zh_cn(url: Optional[str]) -> Optional[str]:
    """导出用：若链接未带 locale=zh-CN 则补上（已有其它 locale 则改为 zh-CN）。"""
    if not url or not isinstance(url, str):
        return url
    s = url.strip()
    if not s:
        return s
    try:
        p = urlparse(s)
        pairs = parse_qsl(p.query, keep_blank_values=True)
        q = {k: v for k, v in pairs}
        if q.get("locale") == "zh-CN":
            return s
        q["locale"] = "zh-CN"
        new_query = urlencode(sorted(q.items()))
        return urlunparse((p.scheme, p.netloc, p.path, p.params, new_query, p.fragment))
    except Exception:
        return s


def _build_row(
    p: Product,
    primary,
    latest_profit: Optional[ProfitRecord],
) -> Dict[str, Any]:
    """单行各字段原始值，key 与导出列 key 一致。"""
    return {
        "id": p.id,
        "tiktok_product_id": p.tiktok_product_id,
        "title": p.title,
        "description": p.description,
        "tiktok_url": ensure_tiktok_url_locale_zh_cn(p.tiktok_url),
        "main_image_url": p.main_image_url,
        "region": p.region,
        "region_label": REGION_LABEL_ZH.get(p.region, p.region),
        "category": p.category,
        "price": _num(p.price),
        "currency": p.currency,
        "price_cny": _num(p.price_cny),
        "original_price": _num(p.original_price),
        "discount": p.discount,
        "sales_volume": p.sales_volume,
        "rating": _num(p.rating),
        "review_count": p.review_count,
        "stock_status": STOCK_LABEL_ZH.get(int(p.stock_status), str(p.stock_status)),
        "shop_name": p.shop_name,
        "shop_id": p.shop_id,
        "seller_location": p.seller_location,
        "shipping_fee": _num(p.shipping_fee),
        "free_shipping": ("是" if int(p.free_shipping) else "否") if p.free_shipping is not None else None,
        "delivery_days_min": p.delivery_days_min,
        "delivery_days_max": p.delivery_days_max,
        "status": p.status,
        "status_label": STATUS_LABEL_ZH.get(str(p.status), str(p.status)),
        "remark": p.remark,
        "estimated_profit": _num(p.estimated_profit),
        "profit_rate": _profit_rate_pct(p.profit_rate),
        "pdd_price": _num(primary.pdd_price) if primary else None,
        "pdd_shop_name": primary.pdd_shop_name if primary else None,
        "pdd_product_url": primary.pdd_product_url if primary else None,
        "pdd_title": primary.pdd_title if primary else None,
        "latest_profit": _num(latest_profit.profit) if latest_profit else None,
        "latest_profit_rate": _profit_rate_pct(latest_profit.profit_rate) if latest_profit else None,
        "created_at": _fmt_dt(p.created_at),
        "updated_at": _fmt_dt(p.updated_at),
    }


# 列顺序与默认导出顺序；key -> 表头中文名
EXPORT_COLUMNS: List[Tuple[str, str]] = [
    ("id", "商品ID"),
    ("tiktok_product_id", "TikTok商品ID"),
    ("title", "商品标题"),
    ("description", "商品描述"),
    ("tiktok_url", "TikTok链接"),
    ("main_image_url", "主图URL"),
    ("region", "区域代码"),
    ("region_label", "区域"),
    ("category", "分类"),
    ("price", "TikTok售价"),
    ("currency", "货币"),
    ("price_cny", "折合人民币"),
    ("original_price", "原价"),
    ("discount", "折扣"),
    ("sales_volume", "销量"),
    ("rating", "评分"),
    ("review_count", "评价数"),
    ("stock_status", "库存状态"),
    ("shop_name", "店铺名称"),
    ("shop_id", "店铺ID"),
    ("seller_location", "卖家地区"),
    ("shipping_fee", "运费"),
    ("free_shipping", "是否包邮"),
    ("delivery_days_min", "配送天数(最少)"),
    ("delivery_days_max", "配送天数(最多)"),
    ("status", "选品状态代码"),
    ("status_label", "选品状态"),
    ("remark", "备注"),
    ("estimated_profit", "预估利润"),
    ("profit_rate", "预估利润率"),
    ("pdd_price", "拼多多价格(主匹配)"),
    ("pdd_shop_name", "拼多多店铺(主匹配)"),
    ("pdd_product_url", "拼多多商品链接(主匹配)"),
    ("pdd_title", "拼多多标题(主匹配)"),
    ("latest_profit", "最新核算利润"),
    ("latest_profit_rate", "最新核算利润率"),
    ("created_at", "添加时间"),
    ("updated_at", "更新时间"),
]

VALID_FIELD_KEYS = frozenset(k for k, _ in EXPORT_COLUMNS)


def export_products_excel(
    db: Session,
    user_id: int,
    fields: Optional[List[str]] = None,
    product_ids: Optional[List[int]] = None,
    export_all: bool = False,
) -> bytes:
    """按所选 ID 或全量与字段导出 Excel；fields 为空则导出全部列。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if not export_all:
        if not product_ids:
            raise ValueError("请选择要导出的商品")

    if fields is None or len(fields) == 0:
        keys = [k for k, _ in EXPORT_COLUMNS]
    else:
        unknown = [f for f in fields if f not in VALID_FIELD_KEYS]
        if unknown:
            raise ValueError(f"未知字段: {', '.join(unknown)}")
        keys = fields

    query = (
        db.query(Product)
        .options(joinedload(Product.pdd_matches))
        .filter(Product.is_deleted == 0)
        .filter(Product.user_id == user_id)
    )
    if not export_all:
        query = query.filter(Product.id.in_(product_ids))
    products = query.order_by(Product.created_at.desc()).all()
    if not products:
        raise ValueError("未找到可导出的商品，请确认所选商品仍属于当前账号")

    if not export_all:
        # 保持与请求 ID 顺序一致（便于对照）
        id_order = {pid: i for i, pid in enumerate(product_ids)}
        products.sort(key=lambda p: id_order.get(p.id, 10**9))

    # 批量取每条商品最新一条利润记录
    latest_by_pid: Dict[int, ProfitRecord] = {}
    if products:
        pids = [p.id for p in products]
        rows = (
            db.query(ProfitRecord)
            .filter(ProfitRecord.product_id.in_(pids))
            .order_by(ProfitRecord.product_id, ProfitRecord.created_at.desc())
            .all()
        )
        for r in rows:
            if r.product_id not in latest_by_pid:
                latest_by_pid[r.product_id] = r

    header_labels = [label for k, label in EXPORT_COLUMNS if k in keys]
    key_list = [k for k, _ in EXPORT_COLUMNS if k in keys]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "选品导出"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(products, 2):
        primary = _get_primary_match(p)
        lp = latest_by_pid.get(p.id)
        row_data = _build_row(p, primary, lp)
        for col_idx, key in enumerate(key_list, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
